from datetime import datetime, timedelta
import json
import openpyxl

from django.contrib import messages
from django.contrib.auth import login, update_session_auth_hash
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils.timezone import make_aware
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
#from .views import asignar_roles_grupo

from .forms import ShiftForm, HrGroupForm
from .models import (
    HrGroup, HrEmployee, AttShift, EmpleadoTurno,
    EmpJob, EmpRole, EmpCostCenter, AttPunch, GrupoSupervisor, PermisoRol, HrGroupJefe,
)

#from API.models import  HrEmployee, EmpRole, PermisoRol
from API.utils.zk_helpers import registrar_en_biometrico
from API.utils.turnos import detectar_turno
from API.utils.reportes import (
    procesar_marcaciones, generar_reporte_basico, reporte_horas_extras
)
from API.utils.permisos import tiene_permiso
from django.utils.timezone import now

from .models import AprobacionHorasExtras


def parse_fecha(fecha_str):
    try:
        return datetime.strptime(fecha_str.strip().rstrip('-'), "%Y-%m-%d").date()
    except Exception:
        return None



# ================== VISTAS GENERALES ==================
def home(request):
    return render(request, 'home.html')


VISTAS_SISTEMA = ['empleados', 'turnos', 'grupos', 'reportes', 'configuraciones']


def configuracion_view(request):
    if not hasattr(request.user, 'hremployee') or request.user.hremployee.emp_role.nombre != "Administrador":
        messages.error(request, "❌ No tienes permiso para ver esta sección.")
        return redirect("home")

    roles_validos = ["Administrador", "Supervisor", "Jefe de Área", "Líder HSE"]
    empleados = HrEmployee.objects.filter(emp_active=True, emp_role__nombre__in=roles_validos)

    roles = EmpRole.objects.all()
    vistas_sistema = ['empleados', 'turnos', 'grupos', 'reportes', 'configuraciones']

    # === GUARDAR PERMISOS SOLO PARA LOS ROLES VISIBLES EN EL HTML ===
    if request.method == "POST" and request.POST.get("guardar_permisos") == "1":
        claves_checkbox = [k for k in request.POST if k.startswith("rol_")]

        claves_recibidas = set()
        for clave in claves_checkbox:
            partes = clave.split("_")
            if len(partes) == 3:
                rol_id = int(partes[1])
                vista = partes[2]
                claves_recibidas.add((rol_id, vista))

    # 🔄 Actualizar solo los permisos recibidos
        for rol_id, vista in claves_recibidas:
            permiso, creado = PermisoRol.objects.get_or_create(
                rol_id=rol_id,
                vista=vista,
                defaults={'tiene_acceso': True}
            )
            if not creado and not permiso.tiene_acceso:
                permiso.tiene_acceso = True
                permiso.save()
                
               
                

    # ⛔ NO TOCAR otros permisos
        messages.success(request, "✅ Permisos actualizados correctamente.")
        return redirect("configuracion")


    # === GUARDAR USUARIOS ===
    if request.method == "POST" and 'crear_usuario' in request.POST:
        usuario_id = request.POST.get("usuario_id")
        username = request.POST.get("username")
        password = request.POST.get("password")
        confirmar_password = request.POST.get("confirmar_password")

        if not all([usuario_id, username, password, confirmar_password]):
            messages.error(request, "❌ Todos los campos son obligatorios.")
            return redirect("configuracion")

        if password != confirmar_password:
            messages.error(request, "⚠️ Las contraseñas no coinciden.")
            return redirect("configuracion")

        try:
            empleado = HrEmployee.objects.get(id=usuario_id)
            if empleado.user:
                empleado.user.username = username
                empleado.user.set_password(password)
                empleado.user.save()
                if request.user == empleado.user:
                    update_session_auth_hash(request, empleado.user)
                    messages.success(request, "🔐 Tu usuario fue restablecido sin cerrar sesión.")
                else:
                    messages.success(request, f"🔐 Usuario actualizado: {empleado.user.username}")
            else:
                if User.objects.filter(username=username).exists():
                    messages.error(request, f"⚠️ El usuario '{username}' ya existe.")
                    return redirect("configuracion")
                user = User.objects.create_user(
                    username=username,
                    password=password,
                    first_name=empleado.emp_firstname,
                    last_name=empleado.emp_lastname,
                    email=empleado.emp_email
                )
                empleado.user = user
                empleado.save()
                messages.success(request, f"✅ Usuario creado correctamente para {username}")
        except HrEmployee.DoesNotExist:
            messages.error(request, "❌ Empleado no encontrado.")
        return redirect("configuracion")

    # === CARGAR PERMISOS ===
    permisos_dict = {}
    for rol in roles:
        permisos = PermisoRol.objects.filter(rol_id=rol.id)
        for permiso in permisos:
            clave = f"rol_{rol.id}_{permiso.vista}"
            permisos_dict[clave] = permiso.tiene_acceso

    return render(request, "configuracion.html", {
        "empleados": empleados,
        "roles": roles,
        "vistas": vistas_sistema,
        "permisos": permisos_dict
    })


    





# bloquer usuario de ingreso de sesion
def toggle_usuario_activo(request):
    if not hasattr(request.user, 'hremployee') or request.user.hremployee.emp_role.nombre != "Administrador":
        return JsonResponse({"success": False, "error": "No autorizado"}, status=403)

    usuario_id = request.POST.get("usuario_id")
    try:
        empleado = HrEmployee.objects.get(id=usuario_id)
        if empleado.user:
            empleado.user.is_active = not empleado.user.is_active
            empleado.user.save()
            return JsonResponse({"success": True, "nuevo_estado": empleado.user.is_active})
        else:
            return JsonResponse({"success": False, "error": "Este empleado no tiene usuario creado."})
    except HrEmployee.DoesNotExist:
        return JsonResponse({"success": False, "error": "Empleado no encontrado"})






def asistencia_view(request):
    return render(request, 'grupos.html')

# ================== LOGIN ==================
def login_view(request):
    if request.method == "POST":
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            messages.success(request, f"Bienvenido, {user.username}!")
            return redirect('home')
    else:
        form = AuthenticationForm()
    return render(request, 'login.html', {'form': form})

def logout_view(request):
    return redirect('login')

# ================== TURNOS ==================
#def turnos_view(request):
    #turnos = AttShift.objects.all().order_by('id')  # o .order_by('start_time') si prefieres por hora
    #return render(request, 'turnos.html', {'turnos': turnos})

def lista_turnos(request):
    
    if not tiene_permiso(request.user, 'turnos'):
        messages.error(request, "❌ No tienes permiso para ver esta vista.")
        return redirect("home")
    
    turnos = AttShift.objects.all().order_by('id')
    
    return render(request, 'turnos.html', {'turnos': turnos})

def crear_turno(request):
    if request.method == "POST":
        data = {key: request.POST.get(key) for key in [
            "shift_name", "start_time", "max_entry_time", "end_time",
            "work_hours", "break_type", "break_minutes", "break_start", "break_end", "status"
        ]}
        nuevo_turno = AttShift.objects.create(**data)
        return JsonResponse({"success": True, "id": nuevo_turno.id, **{k: str(getattr(nuevo_turno, k)) for k in data}})
    return JsonResponse({"success": False, "error": "Método no permitido"}, status=400)



def editar_turno(request, id):
    turno = get_object_or_404(AttShift, id=id)
    
    if request.method == "GET":
        return JsonResponse({
            "id": turno.id,
            "shift_name": turno.shift_name,
            "start_time": turno.start_time.strftime('%H:%M'),
            "max_entry_time": turno.max_entry_time.strftime('%H:%M'),
            "end_time": turno.end_time.strftime('%H:%M'),
            "work_hours": turno.work_hours,
            "break_type": turno.break_type,
            "break_minutes": turno.break_minutes,
            "break_start": turno.break_start.strftime('%H:%M') if turno.break_start else "00:00",
            "break_end": turno.break_end.strftime('%H:%M') if turno.break_end else "00:00",
            "status": turno.status
        })

    elif request.method == "POST":
        form = ShiftForm(request.POST, instance=turno)
        if form.is_valid():
            form.save()
            return JsonResponse({"success": True})
        return JsonResponse({"success": False, "error": form.errors}, status=400)










def cambiar_estado_turno(request):
    if request.method == "POST":
        turno_id = request.POST.get("id")
        nuevo_estado = request.POST.get("estado")
        
        print("🛠 CAMBIO RECIBIDO")
        print(f"➡️ ID: {turno_id}")
        print(f"➡️ ESTADO: {nuevo_estado}")
        
        
        
        try:
            turno = AttShift.objects.get(id=turno_id)
            print(f"✅ Turno encontrado: {turno.shift_name}")
            turno.status = nuevo_estado
            turno.save()
            print(f"💾 Guardado: {turno.status}")
            return JsonResponse({"success": True})
        except AttShift.DoesNotExist:
            print("❌ Turno no encontrado")
            return JsonResponse({"success": False, "error": "Turno no encontrado"})
    print("❌ Método no permitido")    
    return JsonResponse({"success": False, "error": "Método no permitido"})

    






# ================== GRUPOS ==================
def crear_grupo(request):
    if request.method == "POST":
        form = HrGroupForm(request.POST)
        if form.is_valid():
            grupo = form.save()
            return JsonResponse({"success": True, "id": grupo.id, "nombre": grupo.nombre}, status=201)
        return JsonResponse({"success": False, "error": form.errors.as_json()}, status=400)
    return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

def lista_grupos(request):
    grupos = HrGroup.objects.all()
    return render(request, 'grupos.html', {'grupos': grupos})

def obtener_grupos(request):
    if request.method == 'GET':
        grupos = list(HrGroup.objects.values('id', 'nombre'))
        return JsonResponse({'grupos': grupos})
    
    
    

def detalle_grupo(request, id):
    try:
        grupo = HrGroup.objects.get(id=id)
        empleados = HrEmployee.objects.filter(emp_group=grupo)
        supervisores = GrupoSupervisor.objects.filter(grupo=grupo)

        jefes = grupo.jefes_planta.all()
        jefes_data = [
            {"id": jefe.id, "nombre": f"{jefe.emp_firstname} {jefe.emp_lastname}"}
            for jefe in jefes
        ]

        supervisores_data = [
            {"id": sup.supervisor.id, "nombre": f"{sup.supervisor.emp_firstname} {sup.supervisor.emp_lastname}"}
            for sup in supervisores
        ]

        empleados_data = [
            f"{emp.emp_firstname} {emp.emp_lastname}"
            for emp in empleados
        ]

        data = {
            'grupo': grupo.nombre,
            'jefes_planta': jefes_data,
            'supervisores': supervisores_data,
            'empleados': empleados_data,
        }

        return JsonResponse(data)
    except HrGroup.DoesNotExist:
        return JsonResponse({'error': 'Grupo no encontrado'}, status=404)

  
@require_POST
def asignar_roles_grupo(request, id):
    grupo = get_object_or_404(HrGroup, id=id)

    # ✅ Siempre borrar TODOS los jefes anteriores
    HrGroupJefe.objects.filter(grupo=grupo).delete()
    jefes_ids = request.POST.getlist('jefe_planta')

    # ✅ Crear solo si se enviaron jefes válidos
    if jefes_ids and any(jefe_id.strip() for jefe_id in jefes_ids):
        for jefe_id in jefes_ids:
            if jefe_id.strip():
                jefe = get_object_or_404(HrEmployee, id=jefe_id)
                if jefe.emp_role.nombre.lower() == "jefe de área":
                    HrGroupJefe.objects.create(grupo=grupo, jefe=jefe)
                else:
                    return JsonResponse({
                        'error': f"El empleado {jefe.emp_firstname} {jefe.emp_lastname} no es Jefe de Área"
                    }, status=400)

    # ✅ Siempre borrar TODOS los supervisores anteriores
    GrupoSupervisor.objects.filter(grupo=grupo).delete()
    supervisores_ids = request.POST.getlist('supervisores')

    # ✅ Crear solo si se enviaron supervisores válidos
    if supervisores_ids and any(sup_id.strip() for sup_id in supervisores_ids):
        for sup_id in supervisores_ids:
            if sup_id.strip():
                supervisor = get_object_or_404(HrEmployee, id=sup_id)
                if supervisor.emp_role.nombre.lower() == "supervisor":
                    GrupoSupervisor.objects.create(grupo=grupo, supervisor=supervisor)

    return JsonResponse({'success': True})
    
    
    







def grupos_view(request):
   
    if not tiene_permiso(request.user, 'grupos'):
        messages.error(request, "❌ No tienes permiso para ver esta vista.")
        return redirect("home")
    
    grupos = HrGroup.objects.all()

    # ✅ Aquí filtras por roles correctos
    empleados_jefes = HrEmployee.objects.filter(emp_role__nombre__iexact="Jefe de Área", emp_active=True)
    empleados_supervisores = HrEmployee.objects.filter(emp_role__nombre__iexact="Supervisor", emp_active=True)


    return render(request, "grupos.html", {
        "grupos": grupos,
        "empleados_jefes": empleados_jefes,
        "empleados_supervisores": empleados_supervisores,
    })
        
   
    

# ================== EMPLEADOS ==================
def empleados_view(request):
    
    if not tiene_permiso(request.user, 'empleados'):
        messages.error(request, "❌ No tienes permiso para ver esta vista.")
        return redirect("home")
    
    return render(request, 'empleados.html')

def lista_empleados(request):
    
    if not tiene_permiso(request.user, 'empleados'):
        messages.error(request, "❌ No tienes permiso para ver esta vista.")
        return redirect("home")
    
    context = {
        "empleados": HrEmployee.objects.all(),
        "grupos": HrGroup.objects.all(),
        "cargos": EmpJob.objects.all(),
        "roles": EmpRole.objects.all(),
        "centros_costos": EmpCostCenter.objects.all(),
    }
    return render(request, "empleados.html", context)

def crear_empleado(request):
    if request.method == "POST":
        try:
            ip_biometrico = request.POST.get("biometrico")  # 🔹 Ciudad seleccionada
            emp_pin = request.POST["emp_pin"]
            emp_nombre = request.POST["emp_firstname"]
            
            data = {
                "emp_pin": request.POST["emp_pin"],
                "emp_firstname": request.POST["emp_firstname"],
                "emp_lastname": request.POST["emp_lastname"],
                "emp_job": get_object_or_404(EmpJob, id=request.POST["emp_job"]),
                "emp_group": get_object_or_404(HrGroup, id=request.POST["emp_group"]),
                "emp_role": get_object_or_404(EmpRole, id=request.POST["emp_role"]),
                "emp_cost_center": get_object_or_404(EmpCostCenter, id=request.POST["emp_cost_center"]),
                "emp_email": request.POST["emp_email"],
                "emp_active": request.POST.get("emp_active") == "on",
                
            }
            HrEmployee.objects.create(**data)
            
              # 🔹 REGISTRAR EN EL BIOMÉTRICO
            if ip_biometrico:
                registrar_en_biometrico(emp_pin, emp_nombre, ip_biometrico)
            
            
            return redirect("empleados")
        except Exception as e:
            context = {
                "empleados": HrEmployee.objects.all(),
                "grupos": HrGroup.objects.all(),
                "cargos": EmpJob.objects.all(),
                "roles": EmpRole.objects.all(),
                "centros_costos": EmpCostCenter.objects.all(),
                "error": str(e),
            }
            return render(request, "empleados.html", context)
    return redirect("empleados")


# views.py
def cambiar_estado_empleado(request):
    if request.method == "POST":
        empleado_id = request.POST.get("id")
        nuevo_estado = request.POST.get("estado")

        try:
            empleado = HrEmployee.objects.get(id=empleado_id)
            empleado.emp_active = True if nuevo_estado == "Activo" else False
            empleado.save()
            return JsonResponse({"success": True})
        except HrEmployee.DoesNotExist:
            return JsonResponse({"success": False, "error": "Empleado no encontrado"})

    return JsonResponse({"success": False, "error": "Método no permitido"})






def obtener_empleado(request, id):
    empleado = get_object_or_404(HrEmployee, id=id)
    data = {
        "emp_pin": empleado.emp_pin,
        "emp_firstname": empleado.emp_firstname,
        "emp_lastname": empleado.emp_lastname,
        "emp_job": empleado.emp_job.id,
        "emp_group": empleado.emp_group.id,
        "emp_role": empleado.emp_role.id,
        "emp_cost_center": empleado.emp_cost_center.id,
        "emp_email": empleado.emp_email,
        "emp_active": empleado.emp_active,
    }
    return JsonResponse(data)




@csrf_exempt
def editar_empleado(request, id):
    if request.method == "POST":
        try:
            empleado = get_object_or_404(HrEmployee, id=id)

            # Actualiza datos del empleado
            empleado.emp_pin = request.POST.get("emp_pin")
            empleado.emp_firstname = request.POST.get("emp_firstname")
            empleado.emp_lastname = request.POST.get("emp_lastname")
            empleado.emp_email = request.POST.get("emp_email")
            empleado.emp_job_id = request.POST.get("emp_job")
            empleado.emp_group_id = request.POST.get("emp_group")
            empleado.emp_role_id = request.POST.get("emp_role")
            empleado.emp_cost_center_id = request.POST.get("emp_cost_center")
            empleado.emp_active = request.POST.get("emp_active") == "true"
            empleado.save()

            # ⏺️ Verifica si se quiere registrar en el biométrico
            ip_biometrico = request.POST.get("biometrico")
            if ip_biometrico:
                registrar_en_biometrico(empleado.emp_pin, empleado.emp_firstname, ip_biometrico)

            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    return JsonResponse({"error": "Método no permitido"}, status=405)







@csrf_exempt
def crear_cargo(request):
    if request.method == "POST":
        nombre = request.POST.get("nombre")
        if nombre:
            EmpJob.objects.create(nombre=nombre)
            return JsonResponse({"success": True})
    return JsonResponse({"success": False}, status=400)

@csrf_exempt
def crear_rol(request):
    if request.method == "POST":
        nombre = request.POST.get("nombre")
        if nombre:
            EmpRole.objects.create(nombre=nombre)
            return JsonResponse({"success": True})
    return JsonResponse({"success": False}, status=400)

@csrf_exempt
def crear_centro_costo(request):
    if request.method == "POST":
        nombre = request.POST.get("nombre")
        if nombre:
            EmpCostCenter.objects.create(nombre=nombre)
            return JsonResponse({"success": True})
    return JsonResponse({"success": False}, status=400)





# ================== REPORTES ==================
def reporte_turnos(request):
    asignaciones = EmpleadoTurno.objects.select_related('empleado', 'turno').all()
    return render(request, 'reportes/turnos.html', {'asignaciones': asignaciones})




 # asegúrate de tener esto
from datetime import date, timedelta

def reportes_view(request):
    if not tiene_permiso(request.user, 'reportes'):
        messages.error(request, "❌ No tienes permiso para ver esta vista.")
        return redirect("home")

    tipo = request.GET.get("tipo", "basico")
    buscar = request.GET.get("buscar", "").strip()
    grupo = request.GET.get("grupo")
    desde = request.GET.get("desde")
    hasta = request.GET.get("hasta")

    # ✅ Fechas por defecto
    fecha_inicio = parse_fecha(desde) if desde else date.today() - timedelta(days=3)
    fecha_fin = parse_fecha(hasta) if hasta else date.today()

    filtros = {
        'buscar': buscar,
        'grupo': None,
        'desde': fecha_inicio,
        'hasta': fecha_fin,
    }

    grupos_contexto = HrGroup.objects.none()

    if hasattr(request.user, 'hremployee'):
        empleado = request.user.hremployee
        rol = empleado.emp_role.nombre.lower()

        if rol in ['administrador', 'líder hse']:
            grupos_contexto = HrGroup.objects.all()
            if grupo:
                try:
                    filtros['grupo'] = int(grupo)
                except ValueError:
                    return render(request, 'reportes.html', {'error': '❌ Grupo inválido.'})
        else:
            # ✅ Aquí obtenemos todos sus grupos
            grupos_jefe = HrGroup.objects.filter(jefes_planta=empleado)
            grupos_supervisor = HrGroup.objects.filter(gruposupervisor__supervisor=empleado)
            grupos_contexto = (grupos_jefe | grupos_supervisor).distinct()

            if grupo:
                try:
                    grupo_id = int(grupo)
                    if grupo_id not in grupos_contexto.values_list('id', flat=True):
                        return render(request, 'reportes.html', {'error': '❌ No tienes permiso para ver este grupo.'})
                    filtros['grupo'] = grupo_id
                except ValueError:
                    return render(request, 'reportes.html', {'error': '❌ Grupo inválido.'})
            elif grupos_contexto.exists():
                # ✅ Pasar TODOS los IDs al filtro
                filtros['grupo'] = list(grupos_contexto.values_list('id', flat=True))
            else:
                return render(request, 'reportes.html', {'error': '❌ No tienes grupos asignados.'})

    print("🔍 Filtros usados:", filtros)

    # ✅ Llama a la función del reporte
    datos = generar_reporte_basico(filtros) if tipo == 'basico' else reporte_horas_extras(filtros)

    # ✅ Filtro de búsqueda
    if buscar:
        datos = [r for r in datos if
                 buscar.lower() in str(r["cedula"]).lower() or
                 buscar.lower() in r["apellidos"].lower()]

    datos = sorted(datos, key=lambda r: (r["apellidos"].lower(), r["fecha"]))

    return render(request, 'reportes.html', {
        'datos': datos,
        'tipo': tipo,
        'grupos': grupos_contexto
    })







def ejecutar_procesamiento(request):
    hoy = datetime.now()
    hace_tres_dias = hoy - timedelta(days=3)

    FECHA_INICIO = make_aware(hace_tres_dias.replace(hour=0, minute=0, second=0))
    FECHA_FIN = make_aware(hoy.replace(hour=23, minute=59, second=59))

    procesar_marcaciones(FECHA_INICIO, FECHA_FIN)

    return redirect('reportes_view')  # 👈 asegúrate que el name en urls.py sea correcto

@csrf_exempt
def aprobar_horas_extra(request):
    if request.method == "POST":
        data = json.loads(request.body)
        turno_id = data.get("id")
        field = data.get("field")
        value = data.get("value")

        try:
            turno = EmpleadoTurno.objects.get(id=turno_id)
            empleado_actual = request.user.hremployee
            rol = empleado_actual.emp_role.nombre.lower()

            if field not in ["aprobado_supervisor", "aprobado_jefe_area"]:
                return JsonResponse({"success": False, "error": "Campo no válido"})

            # Obtener o crear aprobación
            aprobacion, _ = AprobacionHorasExtras.objects.get_or_create(turno=turno)

            # 🔒 Si ya está aprobado por ambos, no permitir cambios
            if aprobacion.aprobado_supervisor and aprobacion.aprobado_jefe_area:
                return JsonResponse({"success": False, "error": "❌ Ya fue aprobado por ambos. No se puede modificar."})

            # 🔒 Si el jefe ya aprobó, el supervisor NO puede desmarcar
            if rol == "supervisor" and field == "aprobado_supervisor" and not value and aprobacion.aprobado_jefe_area:
                return JsonResponse({"success": False, "error": "❌ No puedes desmarcar. Ya fue aprobado por el jefe."})

            # ✅ Actualizar campo en EmpleadoTurno
            setattr(turno, field, value)
            turno.save()

            # ✅ Registrar datos en AprobacionHorasExtras
            if field == "aprobado_supervisor" and rol in ["supervisor", "jefe de área"]:
                aprobacion.aprobado_supervisor = value
                aprobacion.supervisor_aprobo = empleado_actual
                aprobacion.fecha_aprobacion_supervisor = now()

            elif field == "aprobado_jefe_area" and rol == "jefe de área":
                aprobacion.aprobado_jefe_area = value
                aprobacion.jefe_aprobo = empleado_actual
                aprobacion.fecha_aprobacion_jefe = now()

            # 🔒 Si ambos marcaron, bloquear para pago
            if aprobacion.aprobado_supervisor and aprobacion.aprobado_jefe_area:
                aprobacion.bloqueado_para_pago = True

            aprobacion.save()

            return JsonResponse({"success": True})

        except EmpleadoTurno.DoesNotExist:
            return JsonResponse({"success": False, "error": "❌ Registro no encontrado"})

    return JsonResponse({"success": False, "error": "❌ Método no permitido"}, status=405)







 # Asegúrate de tener esta función

@login_required
def exportar_excel_general(request):
    tabla = request.GET.get("tabla")

    # Diccionario de modelos permitidos
    modelos = {
        "empleados": HrEmployee,
        "grupos": HrGroup,
        "turnos": AttShift,
        "cargos": EmpJob,
        "centros": EmpCostCenter,
        "roles": EmpRole,
    }

    # 🟡 Caso especial: Reportes (básico y horas extras)
    if tabla == "reportes":
        tipo = request.GET.get("tipo", "basico")
        #cedula = request.GET.get("cedula", "")
        #apellidos = request.GET.get("apellidos", "")
        buscar = request.GET.get("buscar", "").strip()
        grupos = request.GET.get("grupos", "")
        desde = parse_fecha(request.GET.get("desde"))
        hasta = parse_fecha(request.GET.get("hasta"))

        filtros = {
            "buscar": buscar,
            #"cedula": cedula,
            #"apellidos": apellidos,
            "grupo": grupos,
            "desde": desde,
            "hasta": hasta,
        }

        datos = reporte_horas_extras(filtros) if tipo == "horas_extras" else generar_reporte_basico(filtros)
        datos = sorted(datos, key=lambda r: (r["apellidos"].lower(), r["fecha"]))
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Reporte {tipo}"

        if tipo == "basico":
            headers = ["Cédula", "Apellido", "Nombre", "Grupo", "Fecha", "Turno"] + [f"Marca {i+1}" for i in range(8)]
        else:
            headers = [
                "Cédula", "Apellido", "Nombre", "Grupo", "Fecha", "Turno", "Entrada", "Salida",
                "HE Diurnas", "HE Nocturnas", "HE Festivo Diurno", "HE Festivo Nocturno",
                "Recargo Nocturno", "Recargo N Festivo", "Aprobado Sup.", "Aprobado Jefe"
            ]
        ws.append(headers)

        for r in datos:
            if tipo == "basico":
                fila = [r["cedula"], r["apellidos"], r["nombre"], r["grupo"], str(r["fecha"]), r["turno"]]
                fila += [e.strftime("%H:%M") if e else "" for e in r["entradas_salidas"]]
            else:
                fila = [
                    r["cedula"], r["apellidos"], r["nombre"], r["grupo"], str(r["fecha"]), r["turno"],
                    r["entrada"], r["salida"],
                    r["horas_extras_diurnas"], r["horas_extras_nocturnas"],
                    r["horas_extras_festivas_diurnas"], r["horas_extras_festivas_nocturnas"],
                    r["recargo_nocturno"], r["recargo_nocturno_festivo"],
                    "✅" if r["aprobado_supervisor"] else "❌",
                    "✅" if r["aprobado_jefe_area"] else "❌",
                ]
            ws.append(fila)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=reportes_{tipo}.xlsx'
        wb.save(response)
        return response

    # 🟡 Caso especial: Usuarios con acceso
    elif tabla == "configuracion":
        empleados = HrEmployee.objects.filter(user__isnull=False).select_related('emp_role', 'user')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Usuarios del Sistema"

        headers = ["Nombre", "Apellido", "Cédula", "Rol", "Usuario", "Activo"]
        ws.append(headers)

        for emp in empleados:
            fila = [
                emp.emp_firstname,
                emp.emp_lastname,
                emp.emp_pin,
                emp.emp_role.nombre if emp.emp_role else "Sin rol",
                emp.user.username if emp.user else "No asignado",
                "Sí" if emp.user and emp.user.is_active else "No"
            ]
            ws.append(fila)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=usuarios_configuracion.xlsx'
        wb.save(response)
        return response

    # 🟡 Caso especial: Empleados (manejo personalizado por campos relacionados)
    elif tabla == "empleados":
        empleados = HrEmployee.objects.all().select_related('emp_job', 'emp_group', 'emp_role', 'emp_cost_center')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Empleados"

        headers = ["Cédula", "Nombre", "Apellido", "Cargo", "Grupo", "Rol", "Centro de Costo", "Correo", "Activo"]
        ws.append(headers)

        for emp in empleados:
            fila = [
                emp.emp_pin,
                emp.emp_firstname,
                emp.emp_lastname,
                emp.emp_job.nombre if emp.emp_job else "Sin cargo",
                emp.emp_group.nombre if emp.emp_group else "Sin grupo",
                emp.emp_role.nombre if emp.emp_role else "Sin rol",
                emp.emp_cost_center.nombre if emp.emp_cost_center else "Sin centro",
                emp.emp_email,
                "Sí" if emp.emp_active else "No"
            ]
            ws.append(fila)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=empleados.xlsx'
        wb.save(response)
        return response

    # 🟥 Validación si no está en el diccionario de modelos
    if tabla not in modelos:
        return HttpResponse("Tabla no válida", status=400)

    # 🟢 Exportación general para otras tablas
    modelo = modelos[tabla]
    queryset = modelo.objects.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{tabla.capitalize()}"

    campos = [field.name for field in modelo._meta.fields]
    ws.append(campos)

    for obj in queryset:
        fila = []
        for campo in campos:
            valor = getattr(obj, campo)
            if hasattr(valor, "__str__") and not isinstance(valor, (int, float, str, bool)):
                fila.append(str(valor))
            else:
                fila.append(valor)
        ws.append(fila)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={tabla}.xlsx'
    wb.save(response)
    return response




def historial_horas_extras_ajax(request):
    aprobaciones = AprobacionHorasExtras.objects.select_related(
        'turno', 'turno__empleado', 'turno__empleado__emp_group',
         'supervisor_aprobo', 'jefe_aprobo'
    ).filter(
        aprobado_supervisor=True,
        aprobado_jefe_area=True,
        #bloqueado_para_pago=False
    ).order_by('turno__empleado__emp_lastname', 'turno__fecha')

    datos = []
    for a in aprobaciones:
        t = a.turno
        datos.append({
            "cedula": t.empleado.emp_pin,
            "apellidos": t.empleado.emp_lastname,
            "nombre": t.empleado.emp_firstname,
            "grupo": t.empleado.emp_group.nombre if t.empleado.emp_group else "",
            "fecha": str(t.fecha),
            "turno": t.turno.shift_name if t.turno else "",

            "entrada": t.hora_entrada.strftime("%H:%M") if t.hora_entrada else "",
            "salida": t.hora_salida.strftime("%H:%M") if t.hora_salida else "",
            "horas_extras_diurnas": t.horas_extras_diurnas,
            "horas_extras_nocturnas": t.horas_extras_nocturnas,
            "horas_extras_festivas_diurnas": t.horas_extras_festivas_diurnas,
            "horas_extras_festivas_nocturnas": t.horas_extras_festivas_nocturnas,
            "recargo_nocturno": t.recargo_nocturno,
            "recargo_nocturno_festivo": t.recargo_nocturno_festivo,
            "supervisor": f"{a.supervisor_aprobo.emp_firstname} {a.supervisor_aprobo.emp_lastname}" if a.supervisor_aprobo else "",
            "jefe": f"{a.jefe_aprobo.emp_firstname} {a.jefe_aprobo.emp_lastname}" if a.jefe_aprobo else "",
            "bloqueado": 1 if a.bloqueado_para_pago else 0,
            "id_turno": t.id,  # ✅ AGREGAR ESTA LÍNEA
        })

    return JsonResponse(datos, safe=False)


@csrf_exempt
def aprobar_horas_extra(request):
    if request.method == "POST":
        data = json.loads(request.body)
        turno_id = data.get("id")
        field = data.get("field")
        value = data.get("value")

        try:
            turno = EmpleadoTurno.objects.get(id=turno_id)
            empleado_actual = request.user.hremployee
            rol = empleado_actual.emp_role.nombre.lower()

            #if field not in ["aprobado_supervisor", "aprobado_jefe_area"]:
            if field not in ["aprobado_supervisor", "aprobado_jefe_area", "bloqueado_para_pago"]:
                return JsonResponse({"success": False, "error": "Campo no válido"})


            aprobacion, _ = AprobacionHorasExtras.objects.get_or_create(turno=turno)

            # 🔒 Si ya fue aprobado por ambos y no es administrador, no permitir cambios
            if (
                aprobacion.aprobado_supervisor and
                aprobacion.aprobado_jefe_area and
                rol != "administrador"
            ):
                return JsonResponse({"success": False, "error": "❌ Ya fue aprobado por ambos. No se puede modificar."})

            # 🔒 Supervisor no puede desmarcar si jefe ya aprobó
            if rol == "supervisor" and field == "aprobado_supervisor" and not value and aprobacion.aprobado_jefe_area:
                return JsonResponse({"success": False, "error": "❌ No puedes desmarcar. Ya fue aprobado por el jefe."})

            # ✅ Si es ADMIN y desmarca, reinicia toda la aprobación
            if rol == "administrador" and not value:
                turno.aprobado_supervisor = False
                turno.aprobado_jefe_area = False
                turno.save()

                aprobacion.aprobado_supervisor = False
                aprobacion.aprobado_jefe_area = False
                aprobacion.bloqueado_para_pago = False
                aprobacion.save()

                return JsonResponse({"success": True, "reiniciado": True})

            # ✅ Guardar aprobación normal
            setattr(turno, field, value)
            turno.save()

            if field == "aprobado_supervisor" and rol in ["supervisor", "jefe de área"]:
                aprobacion.aprobado_supervisor = value
                aprobacion.supervisor_aprobo = empleado_actual
                aprobacion.fecha_aprobacion_supervisor = now()

            elif field == "aprobado_jefe_area" and rol == "jefe de área":
                aprobacion.aprobado_jefe_area = value
                aprobacion.jefe_aprobo = empleado_actual
                aprobacion.fecha_aprobacion_jefe = now()

            if aprobacion.aprobado_supervisor and aprobacion.aprobado_jefe_area:
                aprobacion.bloqueado_para_pago = True

            aprobacion.save()
            if field == "bloqueado_para_pago" and rol == "administrador":
                aprobacion.bloqueado_para_pago = value
                aprobacion.save()
                return JsonResponse({"success": True})


            return JsonResponse({"success": True})

        except EmpleadoTurno.DoesNotExist:
            return JsonResponse({"success": False, "error": "❌ Registro no encontrado"})

    return JsonResponse({"success": False, "error": "❌ Método no permitido"}, status=405)






